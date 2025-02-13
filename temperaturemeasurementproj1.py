import sys
import time
import pandas as pd
import serial
import os
from tkinter import Tk, filedialog
from openpyxl import load_workbook, Workbook

# Initialize the Tkinter root window (hidden)
root = Tk()
root.withdraw()

# Configure the serial port
try:
    ser = serial.Serial(
        port='COM5',
        baudrate=115200,
        parity=serial.PARITY_NONE,
        stopbits=serial.STOPBITS_ONE,
        bytesize=serial.EIGHTBITS,
        timeout=1  # Read every second
    )
except serial.SerialException as e:
    print(f"Error opening serial port: {e}")
    sys.exit(1)

# Global variables
excel_file = None
current_time = -15 # Hardcoded time that increments with each reading

def choose_save_location():
    """Ask user where to save the Excel file."""
    global excel_file
    file_path = filedialog.asksaveasfilename(
        defaultextension='.xlsx',
        filetypes=[("Excel files", "*.xlsx")],
        title="Choose location to save temperature log"
    )
    if file_path:
        excel_file = file_path
        initialize_excel_file()
        start_logging()

def initialize_excel_file():
    """Creates an Excel file with headers if it doesn't exist."""
    if not os.path.exists(excel_file):
        wb = Workbook()
        ws = wb.active
        ws.title = "Temperature Log"
        ws.append(["Time (s)", "Temperature (°C)", "Timestamp"])
        wb.save(excel_file)

def start_logging():
    """Reads serial data, updates time, and logs it to Excel."""
    global current_time

    # Keep workbook open to reduce file access delay
    wb = load_workbook(excel_file)
    ws = wb.active

    try:
        while True:
            line = ser.readline().decode('utf-8', errors='ignore').strip()

            if line:  # If valid data is received
                try:
                    temperature = float(line)  # Convert serial input to float
                    current_time += 1  # Increment hardcoded time counter
                    timestamp = time.strftime("%Y-%m-%d %H:%M:%S")  # Current time

                    print(f"Logging: Time={current_time}, Temp={temperature}")

                    # Append row directly without closing/reopening file
                    ws.append([current_time, temperature, timestamp])
                    wb.save(excel_file)  # Save after each entry

                except ValueError:
                    print(f"Invalid data received: {line}")

    except KeyboardInterrupt:
        print("Logging stopped by user.")
    finally:
        ser.close()
        wb.close()  # Close workbook when done

# Start process
choose_save_location()